from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    path = Path(r"c:/Users/pc hp/OneDrive/ANTARFOOD/GANTT/GANTT Mantencion temporada baja.xlsm")
    wb = load_workbook(path, data_only=False, keep_vba=True)

    def sheet_state(ws) -> str:
        # openpyxl: "visible" | "hidden" | "veryHidden"
        return getattr(ws, "sheet_state", "visible")

    def first_non_empty_cells(ws, max_rows: int = 30, max_cols: int = 30, limit: int = 8):
        found = []
        for r in range(1, min(max_rows, ws.max_row) + 1):
            for c in range(1, min(max_cols, ws.max_column) + 1):
                v = ws.cell(r, c).value
                if v in (None, ""):
                    continue
                # Keep output compact
                if isinstance(v, str):
                    vv = " ".join(v.strip().split())
                    if len(vv) > 80:
                        vv = vv[:77] + "..."
                    v = vv
                found.append((r, c, v))
                if len(found) >= limit:
                    return found
        return found

    def guess_purpose(sheet_name: str, headers: list[str]) -> str:
        n = sheet_name.lower()
        h = " | ".join([x.lower() for x in headers if x])

        if n == "settings":
            return "Configuración del template (parámetros/validaciones/valores auxiliares)."
        if n == "tareas":
            return "Tabla principal de tareas (input manual del usuario; suele alimentar helpers)."
        if n == "timeline":
            return "Hoja de visualización/resumen tipo línea de tiempo (gráfico/tabla derivada)."
        if "gantt_helper" in n:
            return "Tabla helper para Gantt (tareas normalizadas para que el gráfico lea fechas/estado/responsable)."
        if "task_helper" in n:
            return "Tabla helper de tareas (cálculos/columnas auxiliares usadas por el template y/o Gantt)."

        if "task helper" in h and "due date" in h:
            return "Tabla helper con columnas tipo Task/Due Date (probable fuente directa para importar al sistema)."
        if "fecha" in h and ("tarea" in h or "task" in h):
            return "Hoja de tareas (contiene columnas de fechas y nombres de tarea)."
        return "Hoja auxiliar del template (propósito no obvio sin revisar más contenido)."

    def norm(s: object) -> str:
        return str(s or "").strip().lower().replace("\n", " ")

    def detect_headers(ws, max_scan_rows: int = 60):
        """Return (header_row, headers_list) or (None, [])."""
        want_task = {"task helper", "task", "task name", "tarea", "tareas", "actividad", "nombre", "task"}
        want_due = {"due date helper", "due date", "fecha fin", "fin", "termino", "término", "fecha término", "fecha termino"}

        best = (None, [])
        for r in range(1, min(max_scan_rows, ws.max_row) + 1):
            row = [ws.cell(r, c).value for c in range(1, min(60, ws.max_column) + 1)]
            row_n = [norm(x) for x in row]
            if any(x in want_task for x in row_n) and any(x in want_due for x in row_n):
                return (r, row)
            # Special-case common helper layout: headers at row 4
            if r == 4 and any(x in want_task for x in row_n):
                best = (r, row)
        return best

    def pick_col(headers: list[object], candidates: list[str]) -> int | None:
        hn = [norm(h) for h in headers]
        for c in candidates:
            c_n = norm(c)
            for i, h in enumerate(hn, start=1):
                if h == c_n:
                    return i
        return None

    def count_tasks_in_sheet(ws):
        header_row, headers = detect_headers(ws)
        if not header_row or not headers:
            return None

        task_col = pick_col(headers, [
            "Task Helper",
            "TASK",
            "Task",
            "Tarea",
            "Actividad",
            "Nombre",
        ])
        due_col = pick_col(headers, [
            "Due Date Helper",
            "Due Date",
            "Fecha fin",
            "Fin",
            "Termino",
            "Término",
        ])
        if not task_col:
            return None

        count_name = 0
        count_name_due = 0
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, task_col).value
            if name in (None, ""):
                continue
            count_name += 1
            if due_col:
                due = ws.cell(r, due_col).value
                if due not in (None, ""):
                    count_name_due += 1
        return {
            "header_row": header_row,
            "task_col": task_col,
            "due_col": due_col,
            "rows_with_name": count_name,
            "rows_with_name_and_due": count_name_due if due_col else None,
        }

    print("Sheets (name -> state):")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"- {name}: {sheet_state(ws)}")

    visible = [n for n in wb.sheetnames if sheet_state(wb[n]) == "visible"]
    hidden = [n for n in wb.sheetnames if sheet_state(wb[n]) != "visible"]
    print("\nVisible sheets:", visible)
    print("Hidden sheets:", hidden)

    # Focus on the helper table because it's the cleanest source of task records.
    if "Gantt_Helper" in wb.sheetnames:
        ws = wb["Gantt_Helper"]
        header_row = 1
        headers = [ws.cell(header_row, c).value for c in range(1, 26)]
        print("\n=== Gantt_Helper (key sheet) ===")
        for i, h in enumerate(headers, start=1):
            if h not in (None, ""):
                print(f"C{i:02d}: {h}")

        # Guess some useful columns by name fragments.
        def col_idx(fragment: str) -> int | None:
            fragment_l = fragment.lower()
            for i, h in enumerate(headers, start=1):
                if isinstance(h, str) and fragment_l in h.lower():
                    return i
            return None

        cols = {
            "project": col_idx("project"),
            "phase": col_idx("phase"),
            "task": col_idx("task"),
            "assignee": col_idx("assignee"),
            "team": col_idx("team"),
            "status": col_idx("status"),
            "completion": col_idx("completion"),
            "start": col_idx("start date"),
            "due": col_idx("due date"),
        }
        print("\nDetected columns:")
        for k, v in cols.items():
            print(f"- {k}: {v}")

        # Print a sample of task records (rows where 'Task Helper' looks non-empty)
        task_col = cols.get("task")
        if task_col is None:
            print("\nNo task column detected; cannot sample records.")
            return

        rows = []
        for r in range(2, min(ws.max_row, 2000) + 1):
            task_val = ws.cell(r, task_col).value
            if task_val in (None, ""):
                continue
            rows.append(r)

        print(f"\nTask-like rows found (up to 2000): {len(rows)}")
        print("Sample records (first 10):")
        for r in rows[:10]:
            def get(col_name: str):
                c = cols.get(col_name)
                return ws.cell(r, c).value if c else None

            print(
                {
                    "row": r,
                    "project": get("project"),
                    "phase": get("phase"),
                    "task": get("task"),
                    "assignee": get("assignee"),
                    "team": get("team"),
                    "status": get("status"),
                    "completion": get("completion"),
                    "start": get("start"),
                    "due": get("due"),
                }
            )

    # Quick purpose summary for every sheet
    print("\n=== Sheet purpose summary (best-effort) ===")
    for name in wb.sheetnames:
        ws = wb[name]
        state = sheet_state(ws)
        # Take row 1 as "headers" if it looks like headers; keep it simple.
        headers = [ws.cell(1, c).value for c in range(1, min(31, ws.max_column) + 1)]
        headers_s = [h for h in headers if isinstance(h, str) and h.strip()]
        sample = first_non_empty_cells(ws)
        purpose = guess_purpose(name, headers_s)
        print(f"\n- {name} ({state})")
        print(f"  Purpose: {purpose}")
        if headers_s:
            show = headers_s[:10]
            print(f"  Headers sample: {show}")
        if sample:
            print("  First non-empty cells:")
            for r, c, v in sample:
                print(f"    R{r}C{c}: {v}")

    # Task counts per sheet (best-effort)
    print("\n=== Task counts per sheet (best-effort) ===")
    total_any_name = 0
    total_name_due = 0
    for name in wb.sheetnames:
        ws = wb[name]
        state = sheet_state(ws)
        info = count_tasks_in_sheet(ws)
        if not info:
            print(f"- {name} ({state}): no detectable task table")
            continue
        rows_with_name = int(info["rows_with_name"])
        total_any_name += rows_with_name
        if info["rows_with_name_and_due"] is not None:
            rows_name_due = int(info["rows_with_name_and_due"])
            total_name_due += rows_name_due
            print(
                f"- {name} ({state}): {rows_with_name} filas con nombre | {rows_name_due} con nombre+fecha fin "
                f"(header R{info['header_row']}, task C{info['task_col']}, due C{info['due_col']})"
            )
        else:
            print(
                f"- {name} ({state}): {rows_with_name} filas con nombre "
                f"(header R{info['header_row']}, task C{info['task_col']})"
            )

    print(f"\nTOTAL (sumado por hoja, puede incluir duplicados):")
    print(f"- Filas con nombre de tarea: {total_any_name}")
    print(f"- Filas con nombre + fecha fin: {total_name_due}")

    # Nota: antes había un `else` asociado a este `for` (for-else), lo que imprimía
    # un mensaje confuso. No necesitamos fallback aquí porque ya listamos las hojas arriba.


if __name__ == "__main__":
    main()
